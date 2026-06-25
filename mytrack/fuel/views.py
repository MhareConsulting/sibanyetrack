import csv
import io
import json
from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Avg, Count, Sum
from django.forms import modelformset_factory
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date

from mytrack.vehicles.models import Vehicle

from .models import CalibrationPoint, FuelEvent, FuelEventKind, FuelPriceHistory, FuelReading, TankCalibration


@login_required
def event_list(request):
    org = request.user.organisation
    qs = (
        FuelEvent.objects
        .filter(vehicle__organisation=org)
        .select_related('vehicle')
        .order_by('-occurred_at')
    )

    vehicle_filter = request.GET.get('vehicle', '')
    kind_filter    = request.GET.get('kind', '')
    ack_filter     = request.GET.get('ack', '')

    if vehicle_filter:
        qs = qs.filter(vehicle_id=vehicle_filter)
    if kind_filter:
        qs = qs.filter(kind=kind_filter)
    if ack_filter == '0':
        qs = qs.filter(acknowledged=False)
    elif ack_filter == '1':
        qs = qs.filter(acknowledged=True)

    page     = Paginator(qs, 50).get_page(request.GET.get('page', 1))
    vehicles = Vehicle.objects.filter(organisation=org, is_active=True).order_by('registration')

    # Unacknowledged theft/drain count for badge
    unacked = FuelEvent.objects.filter(
        vehicle__organisation=org,
        kind__in=[FuelEventKind.THEFT, FuelEventKind.DRAIN],
        acknowledged=False,
    ).count()

    return render(request, 'fuel/event_list.html', {
        'page_obj':       page,
        'vehicles':       vehicles,
        'kind_choices':   FuelEventKind.choices,
        'vehicle_filter': vehicle_filter,
        'kind_filter':    kind_filter,
        'ack_filter':     ack_filter,
        'unacked_count':  unacked,
    })


@login_required
def acknowledge_event(request, event_id):
    org   = request.user.organisation
    event = get_object_or_404(FuelEvent, pk=event_id, vehicle__organisation=org)
    if request.method == 'POST':
        event.acknowledged = True
        event.notes = request.POST.get('notes', event.notes)
        event.save(update_fields=['acknowledged', 'notes'])
    return redirect('fuel-event-list')


@login_required
def vehicle_fuel_detail(request, vehicle_id):
    org     = request.user.organisation
    vehicle = get_object_or_404(Vehicle, pk=vehicle_id, organisation=org)

    days = int(request.GET.get('days', 7))
    since = timezone.now() - timedelta(days=days)

    readings = (
        FuelReading.objects
        .filter(vehicle=vehicle, device_timestamp__gte=since)
        .order_by('device_timestamp')
        .values('device_timestamp', 'fuel_level_litres')
    )

    chart_labels  = [r['device_timestamp'].isoformat() for r in readings]
    chart_data    = [round(r['fuel_level_litres'], 2) for r in readings]

    events = (
        FuelEvent.objects
        .filter(vehicle=vehicle, occurred_at__gte=since)
        .order_by('-occurred_at')
    )

    latest = FuelReading.objects.filter(vehicle=vehicle).order_by('-device_timestamp').first()

    # ECU 'total fuel used' burn over the window (CAN/OBD vehicles only).
    counter_vals = list(
        FuelReading.objects
        .filter(vehicle=vehicle, device_timestamp__gte=since, total_fuel_used_litres__isnull=False)
        .order_by('device_timestamp')
        .values_list('total_fuel_used_litres', flat=True)
    )
    ecu_burned = (
        round(counter_vals[-1] - counter_vals[0], 1)
        if len(counter_vals) >= 2 and counter_vals[-1] >= counter_vals[0]
        else None
    )

    # Fuel consumed = sum of negative deltas (reading-to-reading drops, excluding theft/drain events)
    theft_times = list(
        FuelEvent.objects.filter(vehicle=vehicle, kind__in=[FuelEventKind.THEFT, FuelEventKind.DRAIN])
        .values_list('occurred_at', flat=True)
    )
    total_consumed = 0.0
    prev_level = None
    for r in FuelReading.objects.filter(vehicle=vehicle, device_timestamp__gte=since).order_by('device_timestamp'):
        if prev_level is not None:
            delta = r.fuel_level_litres - prev_level
            # Only count gradual consumption drops, not theft spikes
            if -20 <= delta < 0 and not any(
                abs((r.device_timestamp - t).total_seconds()) < 600 for t in theft_times
            ):
                total_consumed += abs(delta)
        prev_level = r.fuel_level_litres

    return render(request, 'fuel/vehicle_detail.html', {
        'vehicle':        vehicle,
        'latest':         latest,
        'events':         events,
        'days':           days,
        'chart_labels':   json.dumps(chart_labels),
        'chart_data':     json.dumps(chart_data),
        'total_consumed': round(total_consumed, 1),
        'ecu_burned':     ecu_burned,
        'reading_count':  len(chart_labels),
    })


@login_required
def vehicle_list(request):
    """Overview: all vehicles with their latest fuel reading."""
    org = request.user.organisation
    vehicles = (
        Vehicle.objects
        .filter(organisation=org, is_active=True)
        .order_by('registration')
    )

    from mytrack.fuel.calibration import has_calibration
    rows = []
    for v in vehicles:
        latest = FuelReading.objects.filter(vehicle=v).order_by('-device_timestamp').first()
        unacked = FuelEvent.objects.filter(
            vehicle=v,
            kind__in=[FuelEventKind.THEFT, FuelEventKind.DRAIN],
            acknowledged=False,
        ).count()
        rows.append({'vehicle': v, 'latest': latest, 'unacked': unacked, 'calibrated': has_calibration(v)})

    return render(request, 'fuel/vehicle_list.html', {'rows': rows})


# ── Calibration editor ─────────────────────────────────────────────────────────

CalibrationPointFormSet = modelformset_factory(
    CalibrationPoint,
    fields=('raw_value', 'litres'),
    extra=5,
    can_delete=True,
)


@login_required
def calibration_editor(request, vehicle_id):
    """
    View and edit the strapping table for a single vehicle.
    Handles the calibration header fields (blind areas, notes) and the point formset.
    """
    org     = request.user.organisation
    vehicle = get_object_or_404(Vehicle, pk=vehicle_id, organisation=org)

    calibration, _ = TankCalibration.objects.get_or_create(vehicle=vehicle)

    if request.method == 'POST':
        # Header fields
        try:
            calibration.bottom_blind_litres = float(request.POST.get('bottom_blind_litres', 0) or 0)
            calibration.top_blind_litres    = float(request.POST.get('top_blind_litres', 0) or 0)
        except (ValueError, TypeError):
            pass
        calibration.notes = request.POST.get('notes', '')
        calibration.save()

        # Points formset
        formset = CalibrationPointFormSet(
            request.POST,
            queryset=CalibrationPoint.objects.filter(calibration=calibration),
        )
        if formset.is_valid():
            instances = formset.save(commit=False)
            for instance in instances:
                instance.calibration = calibration
                instance.save()
            for obj in formset.deleted_objects:
                obj.delete()
            messages.success(request, 'Calibration table saved.')
            return redirect('fuel-calibration-editor', vehicle_id=vehicle.pk)
        # Fall through to re-render with errors
    else:
        formset = CalibrationPointFormSet(
            queryset=CalibrationPoint.objects.filter(calibration=calibration),
        )

    point_count  = CalibrationPoint.objects.filter(calibration=calibration).count()
    has_poly     = bool(calibration.poly_coefficients)

    return render(request, 'fuel/calibration_editor.html', {
        'vehicle':     vehicle,
        'calibration': calibration,
        'formset':     formset,
        'point_count': point_count,
        'has_poly':    has_poly,
    })


@login_required
def calibration_import_csv(request, vehicle_id):
    """
    CSV import for calibration points.
    Accepts a two-column CSV: raw_value, litres (with or without a header row).
    Existing points are replaced.
    """
    org     = request.user.organisation
    vehicle = get_object_or_404(Vehicle, pk=vehicle_id, organisation=org)

    if request.method != 'POST':
        return redirect('fuel-calibration-editor', vehicle_id=vehicle.pk)

    csv_file = request.FILES.get('csv_file')
    if not csv_file:
        messages.error(request, 'No file uploaded.')
        return redirect('fuel-calibration-editor', vehicle_id=vehicle.pk)

    calibration, _ = TankCalibration.objects.get_or_create(vehicle=vehicle)

    try:
        decoded = csv_file.read().decode('utf-8-sig')
        reader  = csv.reader(io.StringIO(decoded))
        points  = []
        for i, row in enumerate(reader):
            if len(row) < 2:
                continue
            raw_str, lit_str = row[0].strip(), row[1].strip()
            # Skip header row if present
            try:
                raw_val = float(raw_str)
                litres  = float(lit_str)
            except ValueError:
                if i == 0:
                    continue  # Header row — skip
                messages.error(request, f'Row {i + 1}: could not parse "{raw_str}", "{lit_str}".')
                return redirect('fuel-calibration-editor', vehicle_id=vehicle.pk)
            points.append(CalibrationPoint(calibration=calibration, raw_value=raw_val, litres=litres))

        if len(points) < 2:
            messages.error(request, 'CSV must contain at least 2 valid data rows.')
            return redirect('fuel-calibration-editor', vehicle_id=vehicle.pk)

        CalibrationPoint.objects.filter(calibration=calibration).delete()
        CalibrationPoint.objects.bulk_create(points, ignore_conflicts=True)
        messages.success(request, f'{len(points)} calibration points imported from CSV.')

    except Exception as exc:
        messages.error(request, f'Import failed: {exc}')

    return redirect('fuel-calibration-editor', vehicle_id=vehicle.pk)


@login_required
def calibration_import_excel(request, vehicle_id):
    """
    Parse an Excel file (FUEL column A, SENSOR N VALUE column B), fit a degree-12
    polynomial using numpy, and store the coefficients on the vehicle's TankCalibration.
    """
    org     = request.user.organisation
    vehicle = get_object_or_404(Vehicle, pk=vehicle_id, organisation=org)

    if request.method != 'POST':
        return redirect('fuel-calibration-editor', vehicle_id=vehicle.pk)

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'No file uploaded.')
        return redirect('fuel-calibration-editor', vehicle_id=vehicle.pk)

    try:
        import numpy as np
        from openpyxl import load_workbook

        wb = load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active

        n_values    = []
        fuel_values = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if not row or len(row) < 2:
                continue
            fuel_raw, n_raw = row[0], row[1]
            # Skip header row (non-numeric first cell)
            if i == 0 and not isinstance(fuel_raw, (int, float)):
                continue
            if fuel_raw is None or n_raw is None:
                continue
            try:
                # openpyxl returns numeric cells as numbers; handle European comma-decimal strings
                fuel_val = float(str(fuel_raw).replace(',', '.'))
                n_val    = float(str(n_raw).replace(',', '.'))
            except (ValueError, TypeError):
                continue
            n_values.append(n_val)
            fuel_values.append(fuel_val)

        if len(n_values) < 13:
            messages.error(
                request,
                f'Need at least 13 data points to fit a degree-12 polynomial (got {len(n_values)}).'
            )
            return redirect('fuel-calibration-editor', vehicle_id=vehicle.pk)

        # Fit: x = N-value (sensor), y = litres (fuel)
        degree = min(12, len(n_values) - 1)
        coeffs = np.polyfit(n_values, fuel_values, degree)
        max_n  = float(max(n_values))

        calibration, _ = TankCalibration.objects.get_or_create(vehicle=vehicle)
        calibration.poly_coefficients = [float(c) for c in coeffs]
        calibration.poly_max_n        = max_n
        calibration.save(update_fields=['poly_coefficients', 'poly_max_n', 'updated_at'])

        messages.success(
            request,
            f'Polynomial fitted (degree {degree}, {len(n_values)} points, max N={max_n:.0f}). '
            f'Active for live readings.'
        )

    except Exception as exc:
        messages.error(request, f'Import failed: {exc}')

    return redirect('fuel-calibration-editor', vehicle_id=vehicle.pk)


@login_required
def calibration_import_polynomial(request, vehicle_id):
    """
    Accept a raw polynomial coefficient string in the format produced by the
    external calibration tool:
        c_n,...,c_1,c_0 : max_n_value ; tank_litres
    and store the coefficients directly on the vehicle's TankCalibration.
    """
    org     = request.user.organisation
    vehicle = get_object_or_404(Vehicle, pk=vehicle_id, organisation=org)

    if request.method != 'POST':
        return redirect('fuel-calibration-editor', vehicle_id=vehicle.pk)

    poly_str = request.POST.get('poly_string', '').strip()
    if not poly_str:
        messages.error(request, 'No polynomial string provided.')
        return redirect('fuel-calibration-editor', vehicle_id=vehicle.pk)

    try:
        # Split on ':' to separate coefficients from max_n[;tank_litres]
        coeff_part, rest = poly_str.split(':', 1)
        max_n_str = rest.split(';', 1)[0].strip()

        coeffs = [float(c.strip()) for c in coeff_part.split(',')]
        if len(coeffs) < 2:
            raise ValueError('At least 2 coefficients required.')

        max_n = float(max_n_str)

        calibration, _ = TankCalibration.objects.get_or_create(vehicle=vehicle)
        calibration.poly_coefficients = coeffs
        calibration.poly_max_n        = max_n
        calibration.save(update_fields=['poly_coefficients', 'poly_max_n', 'updated_at'])

        messages.success(
            request,
            f'Polynomial loaded (degree {len(coeffs) - 1}, max N={max_n:.0f}). '
            f'Active for live readings.'
        )

    except Exception as exc:
        messages.error(request, f'Could not parse polynomial string: {exc}')

    return redirect('fuel-calibration-editor', vehicle_id=vehicle.pk)


# ── Fuel price history ─────────────────────────────────────────────────────────

@login_required
def price_history(request):
    org = request.user.organisation
    records = FuelPriceHistory.objects.filter(organisation=org).order_by('-effective_from')
    return render(request, 'fuel/price_history.html', {'records': records})


@login_required
def fetch_prices_now(request):
    if request.method != 'POST':
        return redirect('fuel-price-history')
    from .cron import fetch_and_store_fuel_prices
    result = fetch_and_store_fuel_prices()
    if 'error' in result:
        messages.error(request, f"Fetch failed: {result['error']}")
    else:
        messages.success(request, f"Fuel prices fetched for {result['effective_from']}. {result['orgs_created']} org(s) updated.")
    return redirect('fuel-price-history')


@login_required
def price_add(request):
    if request.method != 'POST':
        return redirect('fuel-price-history')
    org = request.user.organisation
    effective_from = parse_date(request.POST.get('effective_from', ''))
    if not effective_from:
        messages.error(request, 'Effective date is required.')
        return redirect('fuel-price-history')

    def _dec(key):
        val = request.POST.get(key, '').strip()
        if not val:
            return None
        try:
            from decimal import Decimal
            return Decimal(val)
        except Exception:
            return None

    FuelPriceHistory.objects.update_or_create(
        organisation=org,
        effective_from=effective_from,
        defaults={
            'petrol_95_zar':     _dec('petrol_95_zar'),
            'petrol_93_zar':     _dec('petrol_93_zar'),
            'diesel_500ppm_zar': _dec('diesel_500ppm_zar'),
            'diesel_50ppm_zar':  _dec('diesel_50ppm_zar'),
            'source':            'manual',
        },
    )
    messages.success(request, f'Fuel prices saved for {effective_from}.')
    return redirect('fuel-price-history')
