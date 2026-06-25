from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ---- Brand / style helpers -------------------------------------------------
CYAN = "00C8FF"
PURPLE = "8A2BE2"
PERIWINKLE = "EEF0FB"
INK = "0A0A0A"
YELLOW = "FFFF00"
GREYHDR = "C5CAE9"

FONT = "Inter"
ZAR = '"R"#,##0;("R"#,##0);"-"'
ZAR2 = '"R"#,##0.00;("R"#,##0.00);"-"'
PCT = '0.0%'

thin = Side(style="thin", color="C5CAE9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def f(bold=False, color=INK, size=11, italic=False):
    return Font(name=FONT, bold=bold, color=color, size=size, italic=italic)

def fill(c):
    return PatternFill("solid", start_color=c, end_color=c)

def title_cell(ws, cell, text):
    ws[cell] = text
    ws[cell].font = f(bold=True, color="FFFFFF", size=16)
    ws[cell].fill = fill(PURPLE)
    ws[cell].alignment = Alignment(horizontal="left", vertical="center", indent=1)

def section(ws, cell, text):
    ws[cell] = text
    ws[cell].font = f(bold=True, color="FFFFFF", size=11)
    ws[cell].fill = fill(CYAN)
    ws[cell].alignment = Alignment(horizontal="left", vertical="center", indent=1)

def inputcell(c):
    c.fill = fill(YELLOW)
    c.font = f(color="0000FF")
    c.border = border

def linkcell(c):
    c.font = f(color="008000")

def labelcell(c, bold=False):
    c.font = f(bold=bold)

wb = Workbook()

# ===========================================================================
# SHEET 1 — COST INPUTS
# ===========================================================================
ci = wb.active
ci.title = "Cost Inputs"
ci.sheet_view.showGridLines = False
ci.column_dimensions["A"].width = 46
ci.column_dimensions["B"].width = 16
ci.column_dimensions["C"].width = 52

ci.merge_cells("A1:C1")
title_cell(ci, "A1", "myTrack — Cost Inputs")
ci.row_dimensions[1].height = 26
ci["A2"] = "Track it. Protect it. myTrack.   |   EDIT the yellow cells with YOUR real costs. All values below are illustrative placeholders, not real prices."
ci["A2"].font = f(italic=True, color="555555", size=9)
ci.merge_cells("A2:C2")

def ci_row(r, label, val, note, fmt=ZAR):
    ci[f"A{r}"] = label; ci[f"A{r}"].font = f()
    c = ci[f"B{r}"]; c.value = val; inputcell(c); c.number_format = fmt
    ci[f"C{r}"] = note; ci[f"C{r}"].font = f(color="555555", size=9)

# A. Hardware
section(ci, "A4", "A.  HARDWARE — once-off cost per item")
ci["B4"] = "Cost (R)"; ci["B4"].font = f(bold=True, color="FFFFFF"); ci["B4"].fill = fill(CYAN); ci["B4"].alignment = Alignment(horizontal="center")
ci["C4"] = "Notes"; ci["C4"].font = f(bold=True, color="FFFFFF"); ci["C4"].fill = fill(CYAN)
ci_row(5,  "Teltonika FMB GPS tracker", 1200, "Core unit fitted to every vehicle")
ci_row(6,  "CAN adapter (LV-CAN200)", 1800, "For CAN-bus fuel/ECU data")
ci_row(7,  "OBD harness / plug", 350, "Alternative to CAN on light vehicles")
ci_row(8,  "LLS fuel probe (fallback)", 2200, "Tank probe for non-CAN vehicles")
ci_row(9,  "Panic button", 180, "Driver duress alert")
ci_row(10, "Immobiliser / relay cut-off", 450, "Remote engine disable")
ci_row(11, "RFID / iButton driver ID", 320, "Driver identification")
ci_row(12, "Temperature sensor (reefer)", 650, "Cold-chain monitoring")
ci_row(13, "Backup battery", 280, "Survives main-power tamper")
ci_row(14, "SIM card (once-off)", 50, "Connectivity SIM provisioning")
ci_row(15, "Harness / cabling / consumables", 150, "Fitting consumables per vehicle")

# B. Installation & initiation
section(ci, "A17", "B.  INSTALLATION & INITIATION")
ci["B17"].fill = fill(CYAN); ci["C17"].fill = fill(CYAN)
ci_row(18, "Install labour — Basic (plug & play) /vehicle", 250, "OBD plug, no wiring")
ci_row(19, "Install labour — Standard (wired + CAN) /vehicle", 650, "Hard-wired + CAN tap")
ci_row(20, "Install labour — Advanced (probe + calibration) /vehicle", 1400, "Tank probe + calibration")
ci_row(21, "Auto-electrician call-out /visit", 450, "Per site visit if required")
ci_row(22, "Travel cost /km", 8, "Mobile install / remote sites", ZAR2)
ci_row(23, "Account initiation fee (per account)", 3500, "Platform setup, geofences, users, data load, training")

# C. Monthly run cost per unit
section(ci, "A25", "C.  MONTHLY RUN COST — per unit / month")
ci["B25"].fill = fill(CYAN); ci["C25"].fill = fill(CYAN)
ci_row(26, "SIM / data", 35, "Recurring connectivity")
ci_row(27, "Platform hosting allocation", 18, "Server / Azure share per unit")
ci_row(28, "Software licence", 25, "Per-vehicle platform licence")
ci_row(29, "Support & maintenance reserve", 20, "Helpdesk + field support")
ci_row(30, "Notifications (WhatsApp / SMS)", 12, "Driver & alert messaging")
ci_row(31, "Map / routing API allocation", 8, "Map tiles / speed limits / routing")
ci_row(32, "Warranty / replacement reserve", 15, "Hardware failure provision")

# D. Pricing levers
section(ci, "A34", "D.  PRICING LEVERS")
ci["B34"].fill = fill(CYAN); ci["C34"].fill = fill(CYAN)
ci_row(35, "Target gross margin % (on monthly run cost)", 0.45, "Markup applied to monthly cost", PCT)
ci_row(36, "Hardware markup %", 0.30, "Margin applied to hardware cost", PCT)
ci_row(37, "Installation markup %", 0.20, "Margin applied to install labour", PCT)
ci_row(38, "Default contract term (months)", 36, "Used as default in Quote Builder", "0")
ci_row(39, "Volume discount — Small (5–20 units) %", 0.0, "Applied to monthly price", PCT)
ci_row(40, "Volume discount — Medium (21–100 units) %", 0.08, "Applied to monthly price", PCT)
ci_row(41, "Volume discount — Enterprise (100+ units) %", 0.15, "Applied to monthly price", PCT)

# ===========================================================================
# SHEET 2 — QUOTE BUILDER
# ===========================================================================
qb = wb.create_sheet("Quote Builder")
qb.sheet_view.showGridLines = False
widths = {"A":24,"B":7,"C":11,"D":7,"E":12,"F":7,"G":12,"H":9,"I":13,
          "J":13,"K":12,"L":13,"M":12,"N":13,"O":15,"P":11,"Q":15,"R":13,"S":14,"T":14}
for col,w in widths.items():
    qb.column_dimensions[col].width = w

qb.merge_cells("A1:T1")
title_cell(qb, "A1", "myTrack — Quote Builder")
qb.row_dimensions[1].height = 26
qb["A2"] = "Enter your prospect's fleet below (yellow = your inputs). Green cells pull from Cost Inputs. Everything else is calculated."
qb["A2"].font = f(italic=True, color="555555", size=9)
qb.merge_cells("A2:T2")

# Global inputs row 3-4
qb["A3"] = "Hardware billing:"; qb["A3"].font = f(bold=True)
qb["B3"] = "Amortised"; inputcell(qb["B3"]); qb.merge_cells("B3:C3")
qb["D3"] = "Contract term (months):"; qb["D3"].font = f(bold=True); qb.merge_cells("D3:E3")
qb["F3"] = "=+'Cost Inputs'!B38"; inputcell(qb["F3"]); qb["F3"].number_format = "0"

qb["A4"] = "Total units:"; qb["A4"].font = f(bold=True)
qb["B4"] = "=SUM(B7:B12)"; qb["B4"].font = f(bold=True); qb["B4"].number_format = "0"
qb["D4"] = "Fleet tier:"; qb["D4"].font = f(bold=True); qb.merge_cells("D4:E4")
qb["F4"] = '=IF(B4=0,"-",IF(B4<=20,"Small",IF(B4<=100,"Medium","Enterprise")))'; qb["F4"].font = f(bold=True)
qb["H4"] = "Volume discount applied:"; qb["H4"].font = f(bold=True); qb.merge_cells("H4:J4")
qb["K4"] = "=IF(B4<=20,'Cost Inputs'!B39,IF(B4<=100,'Cost Inputs'!B40,'Cost Inputs'!B41))"
linkcell(qb["K4"]); qb["K4"].number_format = PCT; qb["K4"].font = f(bold=True, color="008000")

# Table header row 6
hdrs = ["Vehicle group","Qty","Data source","Panic","Immobiliser","RFID","Temp sensor","Backup batt","Install level",
        "HW cost/unit","Install cost/unit","HW price/unit","Install price/unit","Run cost/unit/mo",
        "Monthly price/unit","Disc %","Monthly /unit (final)","Once-off/unit","Line once-off","Line monthly"]
for i,h in enumerate(hdrs, start=1):
    c = qb.cell(row=6, column=i, value=h)
    c.font = f(bold=True, color="FFFFFF", size=9)
    c.fill = fill(PURPLE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
qb.row_dimensions[6].height = 38

CI = "'Cost Inputs'!"
sample = [
    ("Long-haul trucks", 12, "CAN", "Y","Y","Y","N","Y","Standard"),
    ("Delivery vans",    14, "OBD", "Y","N","Y","N","N","Basic"),
    ("Reefer trucks",     4, "CAN", "Y","Y","Y","Y","Y","Advanced"),
]
first, last = 7, 12
for idx in range(first, last+1):
    r = idx
    s = sample[idx-first] if idx-first < len(sample) else ("", None, "None","N","N","N","N","N","Basic")
    qb.cell(row=r, column=1, value=s[0])
    for col in range(1,10):
        inputcell(qb.cell(row=r, column=col))
    qb.cell(row=r, column=2, value=s[1]); qb.cell(row=r, column=2).number_format = "0"
    qb.cell(row=r, column=3, value=s[2])
    for col,val in zip(range(4,9), s[3:8]):
        qb.cell(row=r, column=col, value=val); qb.cell(row=r, column=col).alignment = Alignment(horizontal="center")
    qb.cell(row=r, column=9, value=s[8])
    # J HW cost/unit
    qb.cell(row=r, column=10, value=(
        f"=IF(B{r}=0,0,{CI}$B$5+{CI}$B$14+{CI}$B$15"
        f'+IF(C{r}="CAN",{CI}$B$6,IF(C{r}="OBD",{CI}$B$7,IF(C{r}="Probe",{CI}$B$8,0)))'
        f'+IF(D{r}="Y",{CI}$B$9,0)+IF(E{r}="Y",{CI}$B$10,0)+IF(F{r}="Y",{CI}$B$11,0)'
        f'+IF(G{r}="Y",{CI}$B$12,0)+IF(H{r}="Y",{CI}$B$13,0))'))
    # K Install cost/unit
    qb.cell(row=r, column=11, value=(
        f'=IF(I{r}="Basic",{CI}$B$18,IF(I{r}="Standard",{CI}$B$19,IF(I{r}="Advanced",{CI}$B$20,0)))'))
    # L HW price/unit
    qb.cell(row=r, column=12, value=f"=J{r}*(1+{CI}$B$36)")
    # M Install price/unit
    qb.cell(row=r, column=13, value=f"=K{r}*(1+{CI}$B$37)")
    # N run cost/unit/mo
    qb.cell(row=r, column=14, value=f"=SUM({CI}$B$26:$B$32)")
    # O monthly price/unit pre-discount
    qb.cell(row=r, column=15, value=f'=N{r}/(1-{CI}$B$35)+IF($B$3="Amortised",L{r}/$F$3,0)')
    # P disc %
    qb.cell(row=r, column=16, value="=$K$4"); qb.cell(row=r, column=16).number_format = PCT
    # Q monthly final
    qb.cell(row=r, column=17, value=f"=O{r}*(1-P{r})")
    # R once-off/unit
    qb.cell(row=r, column=18, value=f'=M{r}+IF($B$3="Amortised",0,L{r})')
    # S line once-off
    qb.cell(row=r, column=19, value=f"=B{r}*R{r}")
    # T line monthly
    qb.cell(row=r, column=20, value=f"=B{r}*Q{r}")
    for col in [10,11,12,13,14,15,17,18,19,20]:
        qb.cell(row=r, column=col).number_format = ZAR
        qb.cell(row=r, column=col).border = border
    qb.cell(row=r, column=16).border = border

# Rollup
qb["A14"] = "QUOTE SUMMARY"; section(qb, "A14", "QUOTE SUMMARY"); qb.merge_cells("A14:C14")
def roll(r, label, formula, fmt=ZAR, bold=False, link=False):
    qb[f"A{r}"] = label; qb[f"A{r}"].font = f(bold=bold)
    c = qb[f"C{r}"]; c.value = formula; c.number_format = fmt
    c.font = f(bold=bold, color=("008000" if link else INK))
    qb.merge_cells(f"A{r}:B{r}")
roll(15, "Total units", "=B4", "0")
roll(16, "Hardware + installation (once-off)", "=SUM(S7:S12)")
roll(17, "Account initiation fee", f"={CI}B23", link=True)
roll(18, "TOTAL ONCE-OFF", "=C16+C17", bold=True)
roll(19, "Total monthly subscription", "=SUM(T7:T12)", bold=True)
roll(20, "Average monthly per unit", "=IF(C15=0,0,C19/C15)")
roll(21, "Contract term (months)", "=F3", "0")
roll(22, "CONTRACT TOTAL VALUE (TCV)", "=C18+C19*C21", bold=True)
qb["C18"].fill = fill(PERIWINKLE); qb["C19"].fill = fill(PERIWINKLE); qb["C22"].fill = fill(PERIWINKLE)
qb.freeze_panes = "A7"

# Data validations
dv_bill = DataValidation(type="list", formula1='"Upfront,Amortised"', allow_blank=False)
dv_src = DataValidation(type="list", formula1='"CAN,OBD,Probe,None"', allow_blank=True)
dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
dv_lvl = DataValidation(type="list", formula1='"Basic,Standard,Advanced"', allow_blank=True)
qb.add_data_validation(dv_bill); qb.add_data_validation(dv_src)
qb.add_data_validation(dv_yn); qb.add_data_validation(dv_lvl)
dv_bill.add("B3")
dv_src.add(f"C{first}:C{last}")
dv_yn.add(f"D{first}:H{last}")
dv_lvl.add(f"I{first}:I{last}")

# ===========================================================================
# SHEET 3 — TIER PRICING (reference rate card)
# ===========================================================================
tp = wb.create_sheet("Tier Pricing")
tp.sheet_view.showGridLines = False
for col,w in {"A":24,"B":18,"C":20,"D":22,"E":24}.items():
    tp.column_dimensions[col].width = w
tp.merge_cells("A1:E1")
title_cell(tp, "A1", "myTrack — Tier Reference Rate Card")
tp.row_dimensions[1].height = 26
tp["A2"] = "Indicative pricing for a STANDARD config (FMB tracker + CAN + panic button, standard install). Driven live from Cost Inputs."
tp["A2"].font = f(italic=True, color="555555", size=9); tp.merge_cells("A2:E2")

# Reference config building blocks
tp["A4"] = "Reference config — per unit"; tp["A4"].font = f(bold=True)
tp["A5"] = "Hardware cost"; tp["B5"] = f"={CI}B5+{CI}B14+{CI}B15+{CI}B6+{CI}B9"; linkcell(tp["B5"])
tp["A6"] = "Hardware price (with markup)"; tp["B6"] = f"=B5*(1+{CI}B36)"
tp["A7"] = "Install price (Standard)"; tp["B7"] = f"={CI}B19*(1+{CI}B37)"
tp["A8"] = "Monthly run cost"; tp["B8"] = f"=SUM({CI}B26:B32)"; linkcell(tp["B8"])
tp["A9"] = "Base monthly price (pre-discount)"; tp["B9"] = f"=B8/(1-{CI}B35)"
tp["A10"] = "Contract term (months)"; tp["B10"] = f"={CI}B38"; tp["B10"].number_format = "0"; linkcell(tp["B10"])
for r in range(5,10):
    tp[f"B{r}"].number_format = ZAR

# Tier table header
hdr = ["Tier","Fleet size","Once-off / unit\n(upfront HW)","Monthly / unit\n(HW upfront)","Monthly / unit\n(HW amortised)"]
for i,h in enumerate(hdr, start=1):
    c = tp.cell(row=12, column=i, value=h)
    c.font = f(bold=True, color="FFFFFF"); c.fill = fill(PURPLE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = border
tp.row_dimensions[12].height = 32

tiers = [("Small","5 – 20 units", f"{CI}B39"),
         ("Medium","21 – 100 units", f"{CI}B40"),
         ("Enterprise","100+ units", f"{CI}B41")]
for j,(name,rng,disc) in enumerate(tiers):
    r = 13+j
    tp.cell(row=r, column=1, value=name).font = f(bold=True)
    tp.cell(row=r, column=2, value=rng).font = f()
    tp.cell(row=r, column=3, value="=$B$6+$B$7")                       # once-off/unit upfront
    tp.cell(row=r, column=4, value=f"=$B$9*(1-{disc})")                # monthly upfront HW
    tp.cell(row=r, column=5, value=f"=($B$9+$B$6/$B$10)*(1-{disc})")   # monthly amortised
    for col in (3,4,5):
        tp.cell(row=r, column=col).number_format = ZAR
        tp.cell(row=r, column=col).border = border
    tp.cell(row=r, column=1).border = border; tp.cell(row=r, column=2).border = border
tp["A17"] = "Discounts are set per tier in Cost Inputs (section D). Amortised column folds hardware into the monthly fee over the contract term."
tp["A17"].font = f(italic=True, color="555555", size=9); tp.merge_cells("A17:E17")

# ===========================================================================
# SHEET 4 — QUOTE SUMMARY (client-facing one-pager)
# ===========================================================================
qs = wb.create_sheet("Quote Summary")
qs.sheet_view.showGridLines = False
for col,w in {"A":4,"B":34,"C":22,"D":4}.items():
    qs.column_dimensions[col].width = w
qs.merge_cells("B2:C2")
title_cell(qs, "B2", "myTrack — Subscription Proposal")
qs.row_dimensions[2].height = 28
qs["B3"] = "Know where your fleet is. Always."; qs["B3"].font = f(italic=True, color=PURPLE); qs.merge_cells("B3:C3")

qs["B5"] = "Prepared for:"; qs["B5"].font = f(bold=True)
qs["C5"] = "[Client name]"; inputcell(qs["C5"])
qs["B6"] = "Date:"; qs["B6"].font = f(bold=True)
qs["C6"] = "[dd/mm/yyyy]"; inputcell(qs["C6"])
qs["B7"] = "Prepared by:"; qs["B7"].font = f(bold=True)
qs["C7"] = "[Consultant]"; inputcell(qs["C7"])

QB = "'Quote Builder'!"
def qs_row(r, label, formula, fmt=ZAR, big=False):
    qs[f"B{r}"] = label; qs[f"B{r}"].font = f(bold=big)
    c = qs[f"C{r}"]; c.value = formula; c.number_format = fmt
    c.font = f(bold=True, color="008000", size=(13 if big else 11))
    c.alignment = Alignment(horizontal="right")

section(qs, "B9", "Your proposal"); qs["C9"].fill = fill(CYAN)
qs_row(10, "Fleet size (units)", f"={QB}C15", "0")
qs_row(11, "Contract term (months)", f"={QB}C21", "0")
qs["B12"] = "";
section(qs, "B13", "Once-off (set-up)"); qs["C13"].fill = fill(CYAN)
qs_row(14, "Hardware + installation", f"={QB}C16")
qs_row(15, "Account initiation fee", f"={QB}C17")
qs_row(16, "Total once-off", f"={QB}C18", big=True)
qs["C16"].fill = fill(PERIWINKLE)
section(qs, "B18", "Monthly subscription"); qs["C18b" if False else "C18"].fill = fill(CYAN)
qs_row(19, "Average per unit / month", f"={QB}C20")
qs_row(20, "Total monthly", f"={QB}C19", big=True)
qs["C20"].fill = fill(PERIWINKLE)
section(qs, "B22", "Contract total value"); qs["C22"].fill = fill(CYAN)
qs_row(23, "TCV (once-off + monthly × term)", f"={QB}C22", big=True)
qs["C23"].fill = fill(PERIWINKLE)
qs["B25"] = "Pricing indicative and subject to site survey. Valid 30 days. E&OE."
qs["B25"].font = f(italic=True, color="555555", size=9); qs.merge_cells("B25:C25")

wb.save("myTrack_Pricing_Toolkit.xlsx")
print("saved")
